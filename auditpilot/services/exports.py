from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from auditpilot.models import ExceptionRecord, ExceptionStatusChoices


HEADER_FONT = Font(bold=True)


def _append_table(worksheet, headers, rows):
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
    for row in rows:
        worksheet.append(row)
    for column in worksheet.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        worksheet.column_dimensions[column[0].column_letter].width = min(max(values, default=10) + 2, 40)


def build_weekly_pack(run):
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = workbook.create_sheet('Run Summary')
    summary_rows = [
        ('Run ID', run.id),
        ('Status', run.status),
        ('As of label', run.as_of_label),
        ('Uploaded by', run.uploaded_by),
        ('Source file', run.source_file.original_name),
        ('SHA-256', run.source_file.sha256),
        ('Rows processed', run.total_rows),
        ('Exceptions', run.total_exceptions),
        ('Warnings', run.total_warnings),
        ('Errors', run.total_errors),
        ('Started', run.started_at.isoformat()),
        ('Completed', run.completed_at.isoformat() if run.completed_at else ''),
    ]
    _append_table(summary_sheet, ['Metric', 'Value'], summary_rows)

    dq_sheet = workbook.create_sheet('DQ Report')
    dq_rows = [
        (finding.sheet_run.sheet_name if finding.sheet_run else '', finding.severity, finding.code, finding.message, str(finding.details_json))
        for finding in run.dq_findings.select_related('sheet_run').all()
    ]
    _append_table(dq_sheet, ['Sheet', 'Severity', 'Code', 'Message', 'Details'], dq_rows)

    new_exception_sheet = workbook.create_sheet('New Exceptions')
    new_exception_rows = [
        (str(exception.exception_id), exception.entity, exception.severity, exception.title, exception.detail, exception.status, exception.disposition, exception.owner_name, exception.due_date.isoformat() if exception.due_date else '')
        for exception in run.exceptions.select_related('control').all()
    ]
    _append_table(new_exception_sheet, ['Exception ID', 'Entity', 'Severity', 'Title', 'Detail', 'Status', 'Disposition', 'Owner', 'Due Date'], new_exception_rows)

    open_sheet = workbook.create_sheet('Open Exceptions')
    open_exception_rows = [
        (str(exception.exception_id), exception.entity, exception.severity, exception.title, exception.status, exception.disposition, exception.owner_name, exception.run_id)
        for exception in ExceptionRecord.objects.exclude(status=ExceptionStatusChoices.CLOSED).order_by('-opened_at')
    ]
    _append_table(open_sheet, ['Exception ID', 'Entity', 'Severity', 'Title', 'Status', 'Disposition', 'Owner', 'Run ID'], open_exception_rows)

    closed_sheet = workbook.create_sheet('Closed This Period')
    period_end = run.completed_at or run.started_at
    closed_exception_rows = [
        (str(exception.exception_id), exception.entity, exception.severity, exception.title, exception.closed_at.isoformat() if exception.closed_at else '', exception.owner_name)
        for exception in ExceptionRecord.objects.filter(closed_at__gte=run.started_at, closed_at__lte=period_end).order_by('-closed_at')
    ]
    _append_table(closed_sheet, ['Exception ID', 'Entity', 'Severity', 'Title', 'Closed At', 'Owner'], closed_exception_rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
