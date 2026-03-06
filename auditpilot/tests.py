import shutil
import tempfile
from io import BytesIO
from types import SimpleNamespace

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from auditpilot.models import AuditRun, ControlCatalog, ExceptionRecord, ExceptionStatusChoices, RecordClassChoices
from auditpilot.services.constants import SOURCE_SPECS
from auditpilot.services.normalize import normalize_sheet
from auditpilot.services.rules import evaluate_age_over_threshold, evaluate_date_order, evaluate_duplicate_key, evaluate_non_negative_metric, evaluate_required_when, ensure_starter_controls
from auditpilot.services.utils import clean_scalar, parse_date_value, parse_decimal_value


def build_row(headers, mapping):
    counts = {}
    values = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        key = f"{header}__{counts[header]}"
        values.append(mapping.get(key, mapping.get(header, '')))
    return values


def build_workbook_bytes(include_jgs_payment_document=True):
    workbook = Workbook()
    soc_sheet = workbook.active
    soc_sheet.title = 'SOC'
    soc_headers = SOURCE_SPECS['SOC']['header_sequence']
    soc_sheet.append(soc_headers)
    soc_base = {
        'Company Code': '3200',
        'Company Name': 'JG Summit Olefins',
        'Fiscal year': 2026,
        'Vendor Code': '5001',
        'Vendor Name': 'Vendor A',
        'Vendor Account Group': 'A/P Trade Local Supplier',
        'Accounting Document Number': 'AD-001',
        'Reference Document': 'REF-001',
        'Plant': 'Batangas',
        'Payment Year': 2026,
        'Payment Document': 'PD-001',
        'Check Number': 'CHK-001',
        'OR Number': '',
        'Document Type': 'Vendor Payment',
        'Payment Method': 'Check',
        'Document Date': '2026-03-01',
        'Baseline Date': '2026-03-02',
        'Posting Date': '2026-03-05',
        'Date to BCD': '2026-03-06',
        'Clearing': '2026-03-07',
        'Payment Release Date': '2026-03-04',
        'Net due date': '2026-03-15',
        'POT': '1',
        'Nature of Transaction': 'Services',
        'Process Group': 'Indirect',
        'AP Status': 'Open',
        'AP Status Description': 'Pending review',
        'Status Date': '2026-03-06',
        'Amount': '1000',
        'Amount in Doc Currency': '1000',
        'Days of Status': '35',
        'AP PCT (Post to Rele': '12',
        'AP TAT (Pos to Clear': '14',
    }
    soc_sheet.append(build_row(soc_headers, soc_base))
    soc_duplicate = dict(soc_base)
    soc_duplicate['Reference Document'] = 'REF-002'
    soc_sheet.append(build_row(soc_headers, soc_duplicate))

    jgs_sheet = workbook.create_sheet('JGS')
    jgs_headers = list(SOURCE_SPECS['JGS']['header_sequence'])
    if not include_jgs_payment_document:
        jgs_headers.remove('Payment Document')
    jgs_sheet.append(jgs_headers)
    jgs_row = {
        'Company Code': '1000',
        'Company Name': 'JG Summit Holdings',
        'Fiscal year': 2026,
        'Vendor Code': '6001',
        'Vendor Name': 'Vendor B',
        'Vendor Account Group': 'Vendor Regular',
        'Vendor Account Group__2': 'Vendor Regular',
        'Process Group': 'Services',
        'Nature of Transaction': 'IT',
        'Reference Document': 'JREF-001',
        'Payment Year': 2026,
        'Terms of Payment': '30D',
        'SAP Document Number': 'SAP-001',
        'Payment Document': 'JPD-001',
        'Document type': 'Vendor Payment',
        'Document type Desc': 'Vendor payment',
        'Payment Method': 'Bank transfer',
        'Check Number': '',
        'Check Status': 1,
        'Document Date': '2026-03-02',
        'Base Date': '2026-03-03',
        'VIM Entered Date': '2026-03-02',
        'SAP Document Posting date': '2026-03-03',
        'Unblocking Date': '2026-03-03',
        'Payment Entry Date': '2026-03-04',
        'SAP Document Clearing Date': '2026-03-06',
        'Payment Readiness Dt': '2026-03-04',
        'Payment Release Date': '2026-03-05',
        'OR Submission Date': '2026-03-05',
        'Net Due date': '2026-03-10',
        'Adjusted Net Due Date': '2026-03-10',
        'Item': '10',
        'Doc Header Text': 'Invoice',
        'Item Text': 'Service fee',
        'Bank Reference': '',
        'Status': 'Open',
        'Status Description': 'Pending treasury',
        'Status Date': '2026-03-05',
        'Amount in Local Curr': '2500',
        'Days (Term 1)': '10',
        'Days (Term 2)': '10',
        'Days (Term 3)': '10',
        'Days of Status': '5',
        'POT': '1',
        'AP TAT': '2',
        'CP TAT': '1',
        'TR TAT': '1',
        'Payment Cycle Time': '8',
        'Total Processing Time': '-1',
    }
    jgs_sheet.append(build_row(jgs_headers, jgs_row))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class UtilityAndRuleTests(TestCase):
    def test_scalar_normalization_helpers(self):
        self.assertIsNone(clean_scalar('PEP/Not assigned'))
        self.assertEqual(parse_date_value('2026-03-01').isoformat(), '2026-03-01')
        self.assertEqual(str(parse_decimal_value('(1,250.50)')), '-1250.50')

    def test_normalization_derives_record_class(self):
        run = AuditRun(status='PENDING')
        headers = SOURCE_SPECS['SOC']['header_sequence']
        dataframe_row = {header: '' for header in headers}
        dataframe_row.update({'Company Code': '3200', 'Vendor Code': 'V1', 'Vendor Name': 'Vendor A', 'Fiscal year': 2026, 'Accounting Document Number': 'A1', 'Payment Document': 'P1', 'Payment Method': 'Check', 'Document Date': '2026-03-01', 'Posting Date': '2026-03-02', 'Net due date': '2026-03-05', 'Status Date': '2026-03-02', 'Amount': '100', 'Process Group': 'Result'})
        import pandas as pd
        records, metrics = normalize_sheet(run, 'SOC', pd.DataFrame([dataframe_row]))
        self.assertEqual(records[0].record_class, RecordClassChoices.DERIVED_RESULT)
        self.assertEqual(metrics['derived_row_count'], 1)

    def test_rule_templates_flag_expected_conditions(self):
        ensure_starter_controls()
        record = SimpleNamespace(
            payment_method='Check',
            payment_release_date=None,
            source_payload={'Check Number': '', 'OR Number': ''},
            record_class='base_transaction',
            record_fingerprint='abc',
            document_date=parse_date_value('2026-03-03'),
            posting_date=parse_date_value('2026-03-02'),
            baseline_date=parse_date_value('2026-03-01'),
            net_due_date=parse_date_value('2026-03-10'),
            days_of_status=parse_decimal_value('45'),
            payment_cycle_time=parse_decimal_value('-1'),
            total_processing_time=parse_decimal_value('400'),
            status_description='Pending',
            status_code='Open',
            source_row_number=2,
            transaction_key='dup-key',
        )
        required_control = ControlCatalog.objects.get(control_id='SOC-CHECK-REFERENCE-REQUIRED')
        self.assertEqual(len(evaluate_required_when(required_control, [record])), 1)
        date_control = ControlCatalog.objects.get(control_id='ALL-LIFECYCLE-DATE-ORDER')
        self.assertGreaterEqual(len(evaluate_date_order(date_control, [record])), 1)
        metric_control = ControlCatalog.objects.get(control_id='ALL-NON-NEGATIVE-METRICS')
        self.assertGreaterEqual(len(evaluate_non_negative_metric(metric_control, [record])), 2)
        aging_control = ControlCatalog.objects.get(control_id='ALL-AGING-BREACH')
        self.assertEqual(len(evaluate_age_over_threshold(aging_control, [record])), 1)
        duplicate_control = ControlCatalog.objects.get(control_id='ALL-DUPLICATE-TRANSACTION-KEY')
        self.assertEqual(len(evaluate_duplicate_key(duplicate_control, [record, record])), 2)


class UploadFlowTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.temp_media = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.temp_media, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.client = Client()

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_upload_run_processes_workbook_and_exports_pack(self):
        workbook = SimpleUploadedFile('weekly.xlsx', build_workbook_bytes(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response = self.client.post(reverse('auditpilot:upload_run'), {'workbook': workbook, 'as_of_label': '2026-W10', 'uploaded_by': 'Analyst'})
        self.assertEqual(response.status_code, 302)
        run = AuditRun.objects.get()
        self.assertEqual(run.status, 'COMPLETED')
        self.assertEqual(run.as_of_label, '2026-W10')
        self.assertEqual(run.sheet_runs.count(), 2)
        self.assertEqual(run.normalized_records.count(), 3)
        self.assertGreater(run.exceptions.count(), 0)
        export_response = self.client.get(reverse('auditpilot:export_run_pack', args=[run.id]))
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertEqual(workbook.sheetnames, ['Run Summary', 'DQ Report', 'New Exceptions', 'Open Exceptions', 'Closed This Period'])

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_missing_required_column_fails_dq_gate(self):
        workbook = SimpleUploadedFile('weekly-bad.xlsx', build_workbook_bytes(include_jgs_payment_document=False), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response = self.client.post(reverse('auditpilot:upload_run'), {'workbook': workbook})
        self.assertEqual(response.status_code, 302)
        run = AuditRun.objects.get()
        self.assertEqual(run.status, 'FAILED_DQ')
        self.assertEqual(run.normalized_records.count(), 0)
        self.assertEqual(run.exceptions.count(), 0)

    @override_settings(MEDIA_ROOT=tempfile.gettempdir())
    def test_same_workbook_hash_reproduces_same_counts_when_reprocessed_cleanly(self):
        payload = build_workbook_bytes()
        workbook_one = SimpleUploadedFile('same.xlsx', payload, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.client.post(reverse('auditpilot:upload_run'), {'workbook': workbook_one})
        first_run = AuditRun.objects.first()
        first_hash = first_run.source_file.sha256
        first_counts = (first_run.normalized_records.count(), first_run.exceptions.count())

        AuditRun.objects.all().delete()
        ExceptionRecord.objects.all().delete()

        workbook_two = SimpleUploadedFile('same.xlsx', payload, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.client.post(reverse('auditpilot:upload_run'), {'workbook': workbook_two})
        second_run = AuditRun.objects.first()
        second_hash = second_run.source_file.sha256
        second_counts = (second_run.normalized_records.count(), second_run.exceptions.count())

        self.assertEqual(first_hash, second_hash)
        self.assertEqual(first_counts, second_counts)

