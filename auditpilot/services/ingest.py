import hashlib
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from django.utils import timezone
from openpyxl import load_workbook

from auditpilot.models import SourceFileArchive
from auditpilot.services.constants import SOURCE_SPECS
from auditpilot.services.utils import normalize_text


@dataclass
class SheetPayload:
    sheet_name: str
    original_headers: list
    internal_headers: list
    dataframe: pd.DataFrame


@dataclass
class WorkbookPayload:
    sheets: dict


def uniquify_headers(headers):
    counts = Counter()
    internal = []
    for index, header in enumerate(headers, start=1):
        text = normalize_text(header) or f'Unnamed Column {index}'
        counts[text] += 1
        internal.append(text if counts[text] == 1 else f'{text}__{counts[text]}')
    return internal


def _combine_header_rows(header_rows):
    max_length = max((len(row) for row in header_rows), default=0)
    combined = []
    for index in range(max_length):
        pieces = []
        for row in header_rows:
            value = normalize_text(row[index]) if index < len(row) else ''
            if value:
                pieces.append(value)
        if not pieces:
            combined.append(f'Unnamed Column {index + 1}')
            continue
        if len(pieces) == 1:
            combined.append(pieces[0])
            continue
        combined.append(f"{pieces[0]} ({pieces[-1]})" if pieces[-1] not in pieces[:-1] else ' '.join(pieces))
    return combined


def store_uploaded_workbook(uploaded_file, as_of_label='', uploaded_by=''):
    digest = hashlib.sha256()
    for chunk in uploaded_file.chunks():
        digest.update(chunk)
    uploaded_file.seek(0)
    archive = SourceFileArchive.objects.create(
        original_name=Path(uploaded_file.name).name,
        sha256=digest.hexdigest(),
        size_bytes=getattr(uploaded_file, 'size', 0),
        uploaded_by=uploaded_by,
        as_of_label=as_of_label,
    )
    timestamp = timezone.now().strftime('%Y/%m/%d')
    stored_name = f"raw_snapshots/{timestamp}/{archive.sha256}_{Path(uploaded_file.name).name}"
    archive.stored_file.save(stored_name, uploaded_file, save=True)
    return archive


def load_workbook_payload(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    payload = {}
    for worksheet in workbook.worksheets:
        spec = SOURCE_SPECS.get(worksheet.title, {})
        rows = worksheet.iter_rows(values_only=True)
        header_row_count = spec.get('header_rows', 1)
        raw_header_rows = []
        try:
            for _ in range(header_row_count):
                raw_header_rows.append(next(rows))
        except StopIteration:
            continue
        original_headers = _combine_header_rows(raw_header_rows)
        internal_headers = uniquify_headers(original_headers)
        records = []
        for row_number, row in enumerate(rows, start=header_row_count + 1):
            values = list(row)
            if len(values) < len(internal_headers):
                values.extend([None] * (len(internal_headers) - len(values)))
            record = {internal_headers[index]: values[index] for index in range(len(internal_headers))}
            record['__source_row_number'] = row_number
            records.append(record)
        dataframe = pd.DataFrame(records)
        payload[worksheet.title] = SheetPayload(
            sheet_name=worksheet.title,
            original_headers=original_headers,
            internal_headers=internal_headers,
            dataframe=dataframe,
        )
    return WorkbookPayload(sheets=payload)
